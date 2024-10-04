import csv
from collections import defaultdict

import config
from testScripts import tradeMaker, testMaker


def prepare_test_table_list():
    table_list = config.table_list
    return table_list

def make_auto_test():
    table_list = prepare_test_table_list()
    confidence_rate_list = config.confidence_rate_list
    dataTupleList = []
    for confidence_rate in confidence_rate_list:
        config.confidence_rate = confidence_rate
        for table in table_list:
            config.decisionType = table
            print(f'Testing {table} with confidence rate {confidence_rate}')
            tradeList = testMaker.doTest()
            winCount = tradeMaker.winningTradeCount(tradeList)
            loseCount = tradeMaker.losingTradeCount(tradeList)
            winRate = 100 * winCount / (winCount + loseCount)
            loseRate = 100 * loseCount / (winCount + loseCount)
            lastMoney = tradeMaker.lastMoney(tradeList)
            dataTuple = (table, winCount, winRate, loseCount, loseRate, lastMoney)
            dataTupleList.append(dataTuple)
    return dataTupleList


def make_set_dict(dataTupleList, confidence_rate_list, table_list):
    # pick first table_count elements from dataTupleList and set them as first SET
    # pick next table_count elements from dataTupleList and set them as second SET
    # there will be confidence_count SETs
    # each SET will have table_count elements

    SetDict = {}
    for i in range(len(confidence_rate_list)):
        SetDict[confidence_rate_list[i]] = dataTupleList[i*len(table_list):(i+1)*len(table_list)]
    return SetDict


def do_and_write_auto_test_standard():
    dataTupleList = make_auto_test()
    confidence_rate_list = config.confidence_rate_list
    table_list = prepare_test_table_list()
    table_count = len(table_list)
    SetDict = make_set_dict(dataTupleList, confidence_rate_list, table_list)

    with open('auto_test_results.csv', 'w') as f:
        # Write headers for each set
        headers = []
        for confidence_rate in confidence_rate_list:
            headers.extend(
                [f'Table ({confidence_rate})', f'Win Count ({confidence_rate})', f'Win Rate ({confidence_rate})',
                 f'Lose Count ({confidence_rate})', f'Lose Rate ({confidence_rate})',
                 f'Last Money ({confidence_rate})'])
        f.write(','.join(headers) + '\n')

        # Write data for each table
        for i in range(table_count):
            row = []
            for confidence_rate in confidence_rate_list:
                dataTuple = SetDict[confidence_rate][i]
                formatted_last_money = f'"{dataTuple[5]:,}"'
                row.extend([dataTuple[0], dataTuple[1], dataTuple[2], dataTuple[3], dataTuple[4], formatted_last_money])
            f.write(','.join(map(str, row)) + '\n')

    print('Auto Test Results are written to auto_test_results.csv')


def make_auto_test_monthly():
    table_list = prepare_test_table_list()
    confidence_rate_list = config.confidence_rate_list
    dataTupleList = []
    for confidence_rate in confidence_rate_list:
        config.confidence_rate = confidence_rate
        for table in table_list:
            config.decisionType = table
            print(f'Testing {table} with confidence rate {confidence_rate}')
            tradeList = testMaker.doTest()
            monthly_result_tuple_list = tradeMaker.monthly_result_tuple_list(tradeList)
            result = {
                'model': table,
                'confidence_rate': confidence_rate,
                'monthly_results': monthly_result_tuple_list
            }
            dataTupleList.append(result)
    return dataTupleList

import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

def write_monthly_results_to_excel(dataTupleList, excel_filename):
    """
    Aylık sonuçları Excel dosyasına yazar ve modelleri dönüşümlü olarak renklendirir.
    İlk sayfada tam verileri, ikinci sayfada sadece 'Rate' değerlerini yazar.

    Args:
        dataTupleList (list): Her bir model ve confidence rate için sonuçları içeren liste.
        excel_filename (str): Yazılacak Excel dosyasının adı.
    """
    # Yeni bir çalışma kitabı oluşturuyoruz
    wb = openpyxl.Workbook()
    ws_full = wb.active
    ws_full.title = "Full Data"

    # dataTupleList'i 'confidence_rate' e göre gruplandırıyoruz
    confidence_rate_groups = defaultdict(list)
    for data in dataTupleList:
        confidence_rate = data['confidence_rate']
        confidence_rate_groups[confidence_rate].append(data)

    # İlk sayfaya tam verileri yazıyoruz
    row_counter = 1  # Satır sayacımız

    for confidence_rate, models_data in confidence_rate_groups.items():
        # Model isimlerini topluyoruz
        model_names = [data['model'] for data in models_data]

        # Modeller için dönüşümlü renkler tanımlıyoruz
        model_colors = ['FFFFFF', 'FFCC99']  # Beyaz ve koyu turuncu
        # 'FFFFFF' beyaz, 'FFCC99' koyu turuncu

        # Modeller için renk eşlemesi yapıyoruz
        model_color_mapping = {}
        for idx, model_name in enumerate(model_names):
            color_index = idx % len(model_colors)
            model_color_mapping[model_name] = model_colors[color_index]

        # Tüm modellerdeki ayları topluyoruz
        all_months = set()
        month_data_per_model = {}
        for data in models_data:
            model_name = data['model']
            monthly_results = data['monthly_results']  # Aylık sonuçlar listesi
            month_data = {}
            for month_result in monthly_results:
                month = month_result[0]  # Ay numarası
                # (Giriş, Çıkış, Rate) değerlerini alıyoruz
                month_data[month] = month_result[1:]
                all_months.add(month)
            month_data_per_model[model_name] = month_data

        # Ayları sıralıyoruz
        sorted_months = sorted(all_months)

        # Başlık satırlarını yazıyoruz
        # İlk satır: 'ConfidenceRate: X', ModelName1, ModelName2, ...
        header_row_1 = ['ConfidenceRate: ' + str(confidence_rate)]
        for model_name in model_names:
            header_row_1 += [model_name, '', '']
        ws_full.append(header_row_1)

        # Model başlıklarını renklendiriyoruz
        col_counter = 2  # 'ConfidenceRate' sütunundan sonra başlıyor
        for idx, model_name in enumerate(model_names):
            model_color = model_color_mapping[model_name]
            for i in range(3):  # Her model için 3 sütun (IN, OUT, Rate)
                cell = ws_full.cell(row=row_counter, column=col_counter)
                cell.fill = PatternFill(start_color=model_color, end_color=model_color, fill_type='solid')
                cell.font = Font(bold=True)
                col_counter += 1
        row_counter += 1

        # İkinci satır: 'Month No', 'IN', 'OUT', 'Rate' * model sayısı
        header_row_2 = ['Month No']
        for _ in model_names:
            header_row_2 += ['IN', 'OUT', 'Rate']
        ws_full.append(header_row_2)

        # İkinci satırın stilini ayarlıyoruz
        for col in range(1, len(header_row_2) + 1):
            cell = ws_full.cell(row=row_counter, column=col)
            cell.font = Font(bold=True)
        row_counter += 1

        # Verileri yazıyoruz
        for idx, month in enumerate(sorted_months):
            row = [str(month) + '. Ay']  # Ay numarası
            for model_name in model_names:
                month_data = month_data_per_model.get(model_name, {})
                if month in month_data:
                    find_entrance_money, find_exit_money, rate_of_return = month_data[month]
                    row += [find_entrance_money, find_exit_money, rate_of_return]
                else:
                    row += ['', '', '']  # Veri yoksa boş bırakıyoruz
            ws_full.append(row)

            # Ay satırını tamamen renklendiriyoruz
            fill_color = 'FFFFFF' if idx % 2 == 0 else 'D3D3D3'  # Beyaz ve açık gri
            for col in range(1, len(row) + 1):
                cell = ws_full.cell(row=row_counter, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            row_counter += 1

        # Boş satır ekliyoruz
        ws_full.append([])
        row_counter += 1

    # Kolon genişliklerini ayarlıyoruz (9 basamaklı bir sayı genişliğinde)
    standard_width = 12  # Yaklaşık 9 basamaklı bir sayı genişliği
    for column_cells in ws_full.columns:
        ws_full.column_dimensions[column_cells[0].column_letter].width = standard_width

    # İkinci sayfaya sadece 'Rate' değerlerini yazıyoruz
    ws_rate_only = wb.create_sheet(title="Rate Only")
    row_counter = 1  # İkinci sayfanın satır sayacı

    for confidence_rate, models_data in confidence_rate_groups.items():
        # Model isimlerini topluyoruz
        model_names = [data['model'] for data in models_data]

        # Modeller için dönüşümlü renkler tanımlıyoruz
        model_colors = ['FFFFFF', 'FFCC99']  # Beyaz ve koyu turuncu

        # Modeller için renk eşlemesi yapıyoruz
        model_color_mapping = {}
        for idx, model_name in enumerate(model_names):
            color_index = idx % len(model_colors)
            model_color_mapping[model_name] = model_colors[color_index]

        # Tüm modellerdeki ayları topluyoruz
        all_months = set()
        month_data_per_model = {}
        for data in models_data:
            model_name = data['model']
            monthly_results = data['monthly_results']  # Aylık sonuçlar listesi
            month_data = {}
            for month_result in monthly_results:
                month = month_result[0]  # Ay numarası
                # Sadece 'Rate' değerini alıyoruz
                rate_of_return = month_result[3]
                month_data[month] = rate_of_return
                all_months.add(month)
            month_data_per_model[model_name] = month_data

        # Ayları sıralıyoruz
        sorted_months = sorted(all_months)

        # Başlık satırlarını yazıyoruz
        # İlk satır: 'ConfidenceRate: X', ModelName1, ModelName2, ...
        header_row_1 = ['ConfidenceRate: ' + str(confidence_rate)]
        for model_name in model_names:
            header_row_1 += [model_name]
        ws_rate_only.append(header_row_1)

        # Model başlıklarını renklendiriyoruz
        col_counter = 2  # 'ConfidenceRate' sütunundan sonra başlıyor
        for idx, model_name in enumerate(model_names):
            model_color = model_color_mapping[model_name]
            cell = ws_rate_only.cell(row=row_counter, column=col_counter)
            cell.fill = PatternFill(start_color=model_color, end_color=model_color, fill_type='solid')
            cell.font = Font(bold=True)
            col_counter += 1
        row_counter += 1

        # İkinci satır: 'Month No', 'Rate' * model sayısı
        header_row_2 = ['Month No']
        for _ in model_names:
            header_row_2 += ['Rate']
        ws_rate_only.append(header_row_2)

        # İkinci satırın stilini ayarlıyoruz
        for col in range(1, len(header_row_2) + 1):
            cell = ws_rate_only.cell(row=row_counter, column=col)
            cell.font = Font(bold=True)
        row_counter += 1

        # Verileri yazıyoruz
        for idx, month in enumerate(sorted_months):
            row = [str(month) + '. Ay']  # Ay numarası
            for model_name in model_names:
                month_data = month_data_per_model.get(model_name, {})
                if month in month_data:
                    rate_of_return = month_data[month]
                    row += [rate_of_return]
                else:
                    row += ['']  # Veri yoksa boş bırakıyoruz
            ws_rate_only.append(row)

            # Ay satırını tamamen renklendiriyoruz
            fill_color = 'FFFFFF' if idx % 2 == 0 else 'D3D3D3'  # Beyaz ve açık gri
            for col in range(1, len(row) + 1):
                cell = ws_rate_only.cell(row=row_counter, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            row_counter += 1

        # Boş satır ekliyoruz
        ws_rate_only.append([])
        row_counter += 1

    # Kolon genişliklerini ayarlıyoruz (Rate Only sayfası)
    standard_width = 12
    for column_cells in ws_rate_only.columns:
        ws_rate_only.column_dimensions[column_cells[0].column_letter].width = standard_width

    # Excel dosyasını kaydediyoruz
    wb.save(excel_filename)
    print(f'Veriler {excel_filename} dosyasına yazıldı.')


def do_and_write_auto_test_monthly():
    dataTupleList = make_auto_test_monthly()
    write_monthly_results_to_excel(dataTupleList, 'auto_test_monthly_results.csv')


def write_auto_test():
    if config.monthly_auto_test_results:
        do_and_write_auto_test_monthly()
    else:
        do_and_write_auto_test_standard()


